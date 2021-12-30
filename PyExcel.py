# Programme permettant d'obtenir diverses valeurs dans un fichier Excel appelé test.xlsx, puis d'ajouter des valeurs à ce dernier et sauvegarder ces nouvelles valeurs dans un autre fichier Excel appelé output.

import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("test.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
column = sheet["a"]
cells = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]

sheet.append([1, 2, 3])
wb.save("output.xlsx")
